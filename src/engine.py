import pandas as pd
import json
import logging
from pathlib import Path
from typing import Dict, Any, List
from openpyxl import load_workbook
from openpyxl.chart import LineChart, BarChart, DoughnutChart, Reference, Series
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import DataBarRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger(__name__)

class FinancialEngine:
    def __init__(self, client_folder_path: str):
        self.client_path = Path(client_folder_path)
        self.config_path = self.client_path / 'config.json'
        
        if not self.config_path.exists():
            raise FileNotFoundError(f"Config missing: {self.config_path}")
            
        with open(self.config_path, 'r') as f:
            self.config: Dict[str, Any] = json.load(f)

    def _get_category(self, description: str) -> str:
        desc_lower = str(description).lower()
        for category, keywords in self.config['categories'].items():
            for keyword in keywords:
                if keyword in desc_lower:
                    return category
        return "Uncategorized"

    def run(self) -> None:
        # 1. Load & Process Data
        input_path = self.client_path / self.config['files']['input']
        output_path = self.client_path / self.config['files']['output']
        
        logger.info(f"[INFO] Processing: {input_path}")
        df = pd.read_csv(input_path)
        
        # Column Mapping
        mapping = self.config['column_mapping']
        df = df.rename(columns={
            mapping['date']: 'Date',
            mapping['description']: 'Description',
            mapping['amount']: 'Amount'
        })
        
        # Cleaning
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
        df = df.dropna(subset=['Date'])
        df['Description'] = df['Description'].fillna("").astype(str).str.strip()
        df['Amount'] = pd.to_numeric(df['Amount'], errors='coerce').fillna(0)
        df['Category'] = df['Description'].apply(self._get_category)
        
        
        # 1. Expense Breakdown (Total)
        expenses = df[df['Category'] != 'Revenue']
        category_summary = expenses.groupby('Category')['Amount'].sum().reset_index()
        category_summary = category_summary.sort_values('Amount', ascending=False)
        
        df['Year'] = df['Date'].dt.year
        df['MonthNum'] = df['Date'].dt.month
        df['MonthName'] = df['Date'].dt.strftime('%b') # Jan, Feb...
        
        yoy_chart_data = df.pivot_table(index='MonthNum', columns='Year', values='Amount', aggfunc='sum').fillna(0)
        month_map = {1:'Jan', 2:'Feb', 3:'Mar', 4:'Apr', 5:'May', 6:'Jun', 
                     7:'Jul', 8:'Aug', 9:'Sep', 10:'Oct', 11:'Nov', 12:'Dec'}
        yoy_chart_data.insert(0, 'Month', yoy_chart_data.index.map(month_map))
        
        logger.info(f"[INFO] Designing Year-over-Year Dashboard...")
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    
                        yoy_chart_data.to_excel(writer, sheet_name='ChartData', index=False)
            
            df.to_excel(writer, sheet_name='Raw Data', index=False)

        self._style_excel(output_path, category_summary, df)
        logger.info("[SUCCESS] Multi-Year Dashboard Generated.")

    def _style_excel(self, file_path: Path, cat_summary: pd.DataFrame, full_df: pd.DataFrame) -> None:
        wb = load_workbook(file_path)

        if 'Dashboard' not in wb.sheetnames:
            wb.create_sheet('Dashboard', 0)
            
        ws = wb['Dashboard']
        if 'ChartData' in wb.sheetnames:
            ws_chart = wb['ChartData']
        else:
             ws_chart = wb.create_sheet('ChartData')

        ws.sheet_view.showGridLines = False 

        PRIMARY_COLOR = "1F4E78" # Royal Blue
        SECONDARY_COLOR = "D9E1F2" # Light Blue Grey
        WHITE = "FFFFFF"
        
        header_font = Font(name='Calibri', size=11, bold=True, color=WHITE)
        header_fill = PatternFill(start_color=PRIMARY_COLOR, end_color=PRIMARY_COLOR, fill_type='solid')
        
        ws['B2'] = self.config.get('report_title', 'EXECUTIVE FINANCIAL REPORT').upper()
        ws['B2'].font = Font(name='Calibri', size=24, bold=True, color=PRIMARY_COLOR)
        
        ws['B3'] = f"Generated for: {self.client_path.name}"
        ws['B3'].font = Font(size=12, italic=True, color="595959")

        total_rev = full_df[full_df['Category'] == 'Revenue']['Amount'].sum()
        total_exp = full_df[full_df['Category'] != 'Revenue']['Amount'].sum()
        net_profit = total_rev - total_exp
        
        kpis = [("TOTAL REVENUE", total_rev), ("TOTAL EXPENSES", total_exp), ("NET PROFIT", net_profit)]
        
        start_col = 2 # B
        for title, value in kpis:
            cell_top = ws.cell(row=5, column=start_col)
            cell_val = ws.cell(row=6, column=start_col)
            
            cell_top.value = title
            cell_val.value = value
            
            # Style
            cell_top.font = Font(bold=True, color="595959", size=10)
            cell_top.alignment = Alignment(horizontal='left')
            
            cell_val.font = Font(bold=True, size=16, color=PRIMARY_COLOR)
            cell_val.number_format = '"$"#,##0.00_-'
            cell_val.alignment = Alignment(horizontal='left')
            
            thin_border = Side(border_style="thin", color=PRIMARY_COLOR)
            cell_val.border = Border(bottom=thin_border)
            
            start_col += 2 

        table_start_row = 9
        ws.cell(row=table_start_row, column=2, value="EXPENSE BREAKDOWN").font = Font(bold=True, size=12, color=PRIMARY_COLOR)
        
        # Manual Table Construction for full control
        # Header
        header_row = table_start_row + 1
        headers = ["Category", "Amount", "% of Total"]
        for col_idx, h in enumerate(headers, 2):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.value = h
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            
        current_row = header_row + 1
        for _, row in cat_summary.iterrows():
            ws.cell(row=current_row, column=2, value=row['Category']).alignment = Alignment(horizontal='left')
            c_amt = ws.cell(row=current_row, column=3, value=row['Amount'])
            c_amt.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
            
            pct = row['Amount'] / total_exp if total_exp else 0
            c_pct = ws.cell(row=current_row, column=4, value=pct)
            c_pct.number_format = '0.0%'
            c_pct.alignment = Alignment(horizontal='center')
            
            if current_row % 2 == 1:
                for c in range(2, 5):
                    ws.cell(row=current_row, column=c).fill = PatternFill(start_color=SECONDARY_COLOR, end_color=SECONDARY_COLOR, fill_type='solid')

            current_row += 1

        chart1 = DoughnutChart()
        chart1.title = "Expense Distribution"
        chart1.style = 10 
        
        cats = Reference(ws, min_col=2, min_row=header_row+1, max_row=current_row-1)
        data = Reference(ws, min_col=3, min_row=header_row+1, max_row=current_row-1)
        chart1.add_data(data, titles_from_data=False)
        chart1.set_categories(cats)
        
        ws.add_chart(chart1, "F9")

        yearly_start_row = current_row + 4
        ws.cell(row=yearly_start_row, column=2, value="YEARLY TRENDS").font = Font(bold=True, size=12, color=PRIMARY_COLOR)
        
        chart2 = LineChart()
        chart2.title = "Revenue Growth (YoY)"
        chart2.style = 12
        chart2.y_axis.title = "Revenue"
        chart2.height = 10
        chart2.width = 28
        

        if ws_chart.max_row > 1:
            max_col = ws_chart.max_column
            max_row = ws_chart.max_row
            
            cats = Reference(ws_chart, min_col=1, min_row=2, max_row=max_row)
            data = Reference(ws_chart, min_col=2, min_row=1, max_row=max_row, max_col=max_col)
            
            chart2.add_data(data, titles_from_data=True)
            chart2.set_categories(cats)
            
            ws.add_chart(chart2, f"B{yearly_start_row+2}")

        ws.column_dimensions['A'].width = 2
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 5
        
        wb.save(file_path)